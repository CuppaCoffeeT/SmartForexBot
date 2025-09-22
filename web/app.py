from __future__ import annotations
from flask import Flask, jsonify, render_template, request
from flask_cors import CORS
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import traceback
import os, re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook


from api.main import (
    detect_intent, extract_symbol, get_pair_insight, rank_recommendations, explain_signal,
    compare_pairs, extract_currency, best_for_currency, explain_recommendation,
    explain_signal_simple, explain_recommendation_simple, get_overview_snapshot,
    explain_followup_reason, explain_risk_guidance, explain_confidence_detail,
)

app = Flask(__name__, static_folder="static", template_folder="templates")
CORS(app)
_analyzer = SentimentIntensityAnalyzer()
LOG_PATH = Path("logs/chat_log.xlsx")


def log_chat_event(message: str, intent: str, reply: str, symbol: str | None, interval: str, sentiment: str) -> None:
    try:
        LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        if LOG_PATH.exists():
            wb = load_workbook(LOG_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["timestamp", "message", "intent", "symbol", "interval", "sentiment", "reply"])
        ws.append([
            datetime.utcnow().isoformat(),
            message,
            intent,
            symbol or "",
            interval,
            sentiment,
            reply,
        ])
        wb.save(LOG_PATH)
    except Exception:
        # Never let logging errors break the chat flow
        traceback.print_exc()

@app.route("/")
def index():
    return render_template("index.html")

@app.post("/api/chat")
def chat():
    try:
        data = request.get_json(force=True)
        msg: str = data.get("message", "").strip()
        payload_symbol: str | None = data.get("symbol")
        followup = bool(data.get("followup"))
        feature_mode: str = (data.get("feature_mode") or "fib").lower()
        text_symbol: str | None = extract_symbol(msg)
        user_symbol: str | None = text_symbol if text_symbol else (payload_symbol if followup else None)
        sent = _analyzer.polarity_scores(msg)
        sentiment = "positive" if sent["compound"] > 0.2 else ("negative" if sent["compound"] < -0.2 else "neutral")
        interval = (data.get("interval") or "1h").lower()

        intent = detect_intent(msg)

        if intent in {"ask_followup_why", "ask_risk", "ask_confidence"}:
            if not user_symbol:
                reply = "Tell me which pair you’re referring to so I can explain further."
                log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
                return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})
            info = get_pair_insight(user_symbol, interval=interval, feature_mode=feature_mode)
            if intent == "ask_followup_why":
                reply = explain_followup_reason(info)
            elif intent == "ask_risk":
                reply = explain_risk_guidance(info)
            else:
                reply = explain_confidence_detail(info)
            log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
            return jsonify({
                "reply": reply,
                "intent": intent,
                "sentiment": sentiment,
                "symbol": user_symbol,
                "insight": info.get("insight"),
                "levels": info.get("levels"),
                "ohlc": info.get("ohlc"),
            })

        if intent == "ask_topk":
            m = re.search(r"top\s*(\d+)", msg.lower())
            k = int(m.group(1)) if m else 3
            recos = rank_recommendations(topk=min(max(k,1), 10), feature_mode=feature_mode)
            if not recos:
                reply = "No strong retracement setups right now."
                log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
                return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})
            reply = "Top setups:\n" + "\n".join([f"{i+1}. {r['symbol']} → {r['direction']} (conf {r['confidence']})" for i, r in enumerate(recos)])
            log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
            return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment, "recommendations": recos})

        # COMPARE: try to extract two symbols
        if intent == "ask_compare":
            tokens = re.findall(r"\b([A-Za-z]{3})[\/\s-]?([A-Za-z]{3})\b", msg)
            if len(tokens) >= 2:
                a = (tokens[0][0] + tokens[0][1]).upper()
                b = (tokens[1][0] + tokens[1][1]).upper()
                sym_a = extract_symbol(a) or extract_symbol(tokens[0][0] + "/" + tokens[0][1])
                sym_b = extract_symbol(b) or extract_symbol(tokens[1][0] + "/" + tokens[1][1])
                if sym_a and sym_b:
                    cmpres = compare_pairs(sym_a, sym_b)
                    reply = f"{cmpres['better']} looks stronger now — {cmpres['reason']}."
                    log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
                    return jsonify({
                        "reply": reply, "intent": intent, "sentiment": sentiment,
                        "compare": {"a": cmpres["a"], "b": cmpres["b"], "scores": cmpres["scores"], "better": cmpres["better"]}
                    })
            # fallback
            reply = "Tell me two pairs to compare, e.g., 'compare EURUSD vs GBPUSD'."
            log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
            return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})


        if intent == "help":
            reply = (
                "I can: (1) recommend pairs near Fib retracements, (2) analyze a specific pair (e.g., 'EURUSD'), "
                "(3) explain a current signal, and (4) show a chart. Try: 'what’s the best trade now?' or 'EURUSD chart'."
            )
            log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
            return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})

        if intent in {"ask_reco", "ask_chart"} and not user_symbol:
            # If user mentioned a single currency like eur find best pairs containing it
            cur = extract_currency(msg)
            if cur:
                recos = best_for_currency(cur, feature_mode=feature_mode)
                if recos:
                    best = recos[0]
                    try:
                        info = get_pair_insight(best['symbol'], interval=interval, feature_mode=feature_mode)
                    except Exception:
                        info = {"symbol": best['symbol'], "insight": None, "levels": None, "ohlc": None}
                    # Beginner friendly reply
                    reply = explain_recommendation_simple({**info, "symbol": best['symbol']})
                    log_chat_event(msg, intent, reply, best['symbol'], interval, sentiment)
                    return jsonify({
                        "reply": reply, "intent": intent, "sentiment": sentiment,
                        "recommendations": recos,
                        "symbol": best['symbol'],
                        "insight": info.get("insight"),
                        "levels": info.get("levels"),
                        "ohlc": info.get("ohlc"),
                    })
            recos = rank_recommendations(feature_mode=feature_mode)
            if not recos:
                reply = "No strong retracement setups right now."
                log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
                return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})
            best = recos[0]
            try:
                info = get_pair_insight(best['symbol'], interval=interval, feature_mode=feature_mode)
            except Exception:
                info = {"symbol": best['symbol'], "insight": None, "levels": None, "ohlc": None}
            # Beginner friendly reply
            reply = explain_recommendation_simple({**info, "symbol": best['symbol']})
            log_chat_event(msg, intent, reply, best['symbol'], interval, sentiment)
            return jsonify({
                "reply": reply, "intent": intent, "sentiment": sentiment,
                "recommendations": recos,
                "symbol": best['symbol'],
                "insight": info.get("insight"),
                "levels": info.get("levels"),
                "ohlc": info.get("ohlc"),
            })

        # pair-oriented flows
        if user_symbol:
            info = get_pair_insight(user_symbol, interval=interval, feature_mode=feature_mode)
            if intent == "ask_explain":
                reply = explain_signal_simple(info)
            elif intent == "ask_chart":
                reply = f"Showing the current {user_symbol} chart with labelled Fib lines."
            else:
                # default pair status (simple explanation)
                reply = explain_signal_simple(info)
            log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
            return jsonify({
                "reply": reply,
                "intent": intent,
                "sentiment": sentiment,
                "symbol": user_symbol,
                "insight": info.get("insight"),
                "levels": info.get("levels"),
                "ohlc": info.get("ohlc"),
            })

        # Fallback
        reply = "Tell me a pair, e.g., EURUSD, or ask for recommendations."
        log_chat_event(msg, intent, reply, user_symbol, interval, sentiment)
        return jsonify({"reply": reply, "intent": intent, "sentiment": sentiment})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.get("/api/recommendations")
def api_recos():
    feature_mode = (request.args.get("feature_mode") or "fib").lower()
    recos = rank_recommendations(feature_mode=feature_mode)
    return jsonify({"recommendations": recos})


@app.get("/api/overview")
def api_overview():
    interval = (request.args.get("interval") or "1h").lower()
    feature_mode = (request.args.get("feature_mode") or "fib").lower()
    snapshot = get_overview_snapshot(interval=interval, feature_mode=feature_mode)
    return jsonify({"pairs": snapshot})

@app.get("/api/chart")
def api_chart():
    symbol = request.args.get("symbol", "EURUSD=X")
    interval = (request.args.get("interval") or "1h").lower()
    feature_mode = (request.args.get("feature_mode") or "fib").lower()
    info = get_pair_insight(symbol, interval=interval, feature_mode=feature_mode)
    return jsonify({"symbol": symbol, "levels": info["levels"], "ohlc": info["ohlc"]})


@app.get("/api/models")
def models():
    files = [f for f in os.listdir("artifacts") if re.match(r".+_1h(?:_.+)?\.joblib$", f)]
    pairs: list[str] = []
    for f in files:
        base = f
        base = re.sub(r"_1h(?:_.+)?\.joblib$", "", base)
        # Artifacts are saved as SYMBOLX_1h.joblib
        if base.endswith("X"):
            sym = base[:-1] + "=X"
        else:
            # Fallback: assume it's already the 6-letter symbol
            sym = base + "=X"
        pairs.append(sym)
    pairs.sort()
    return jsonify({"pairs": pairs})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5500, debug=True)
